"""Seed database from Ares Consult Excel file (Март sheet)."""
import sys, re, requests

API = "http://127.0.0.1:8000/api"

# --- Parse brand field to extract client name, trademark, doc_type ---
def parse_brand(brand: str):
    brand = brand.strip()

    # Determine doc_type from prefix
    doc_type = "CC"
    if brand.startswith("ДС "):
        doc_type = "DC"
        brand = brand[3:]
    elif brand.startswith("СС "):
        doc_type = "CC"
        brand = brand[3:]
    elif brand.startswith("CC "):
        doc_type = "CC"
        brand = brand[3:]
    elif brand.startswith("СС/ДС ") or brand.startswith("ДС/СС "):
        doc_type = "CC"  # default to CC for mixed
        brand = brand[6:]

    # Remove layer info like "61/2сл", "62/3сл", "62 "
    brand = re.sub(r'^\d{2}/?\d*сл\s*', '', brand)
    brand = re.sub(r'^\d{2}\s+', '', brand)

    # Extract trademark from parentheses
    trademark = None
    m = re.search(r'\(([^)]+)\)', brand)
    if m:
        trademark = m.group(1).strip()
        client_name = re.sub(r'\s*\([^)]+\)\s*', ' ', brand).strip()
    else:
        client_name = brand.strip()

    # Clean up client name
    client_name = re.sub(r'\s+', ' ', client_name).strip()

    return client_name, trademark, doc_type


def parse_tr_ts(tr_ts_raw: str):
    """Convert '007 дет' -> '007/2011', '017 взр' -> '017/2011'"""
    if not tr_ts_raw:
        return "007/2011"  # default
    if "007" in tr_ts_raw:
        return "007/2011"
    if "017" in tr_ts_raw:
        return "017/2011"
    return "007/2011"


def map_status(old_status: str) -> str:
    """Map Excel statuses to new kanban statuses."""
    s = old_status.strip().lower()
    if "1 очередь" in s:
        return "Новый"
    if "2 очередь" in s:
        return "Новый"
    if "выпущен" in s:
        return "Выпущен"
    if "в работе" in s:
        return "В работе"
    if "готов" in s:
        return "Готов"
    return "Новый"


def map_layout(v: str) -> str:
    if not v or v == "-":
        return "Нет"
    if v in ("Готов", "Утвержден", "В процессе", "Нет"):
        return v
    return "Нет"


def map_sample(v: str) -> str:
    if not v or v == "-":
        return "Нет"
    if v == "Получен":
        return "Получен"
    if v == "Нет":
        return "Нет"
    # Person's name means sample is with someone = "Получен"
    return "Получен"


def map_pi(v: str) -> str:
    if not v or v == "-":
        return "Нет"
    if v == "Получен":
        return "Получен"
    if v == "Нет":
        return "Нет"
    return "Нет"


def create_client(name: str) -> int:
    """Create client via API, return id."""
    # Determine company_type
    company_type = "ИП"
    if "ОсОО" in name:
        company_type = "ОсОО"

    payload = {
        "country_type": "KR",
        "company_type": company_type,
        "company_name": name,
        "fio": name.replace("ИП ", "").replace("ОсОО ", "").replace('"', ''),
        "inn": "00000000000000",
    }
    r = requests.post(f"{API}/clients/", json=payload)
    r.raise_for_status()
    return r.json()["id"]


def create_product(trademark: str, weaving: str, layer: int, doc_type: str, tr_ts: str) -> int:
    """Create a minimal product, return id."""
    payload = {
        "article": f"ART-{trademark or 'STD'}"[:20],
        "name": f"Изделие {trademark or 'стандарт'}",
        "product_type": "Платье",  # generic
        "weaving_type": weaving or "Швейка",
        "target_group": "Детская" if "007" in tr_ts else "Взрослая",
        "layer": layer or 2,
        "trademark": trademark,
        "compositions": [{"material_name": "Хлопок", "percentage": 100}],
    }
    r = requests.post(f"{API}/products/", json=payload)
    r.raise_for_status()
    return r.json()["id"]


def main():
    import openpyxl
    wb = openpyxl.load_workbook(
        'C:/Users/Meta Hall/Downloads/Копия Ares Consult бухгалтерия.xlsx',
        data_only=True
    )
    ws = wb['Март']

    created = 0
    client_cache = {}

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        brand_raw = str(row[0]).strip() if row[0] else ""
        if not brand_raw or brand_raw == "Итого":
            continue

        client_name, trademark, doc_type = parse_brand(brand_raw)
        tr_ts_raw = str(row[1]).strip() if row[1] else ""
        tr_ts = parse_tr_ts(tr_ts_raw)
        layer = int(row[2]) if row[2] and str(row[2]).isdigit() else (int(float(row[2])) if row[2] and row[2] != '?' else 2)
        weaving = str(row[3]).strip() if row[3] else "Швейка"
        status = map_status(str(row[4]).strip() if row[4] else "")
        layout = map_layout(str(row[5]).strip() if row[5] else "")
        sample = map_sample(str(row[6]).strip() if row[6] else "")
        cert_body = str(row[7]).strip() if row[7] and str(row[7]).strip() != "-" else None

        expected_date = None
        if row[8]:
            try:
                expected_date = str(row[8])[:10]
            except:
                pass

        actual_date = None
        if row[10]:
            try:
                actual_date = str(row[10])[:10]
            except:
                pass

        original = str(row[11]).strip() if row[11] else "Нет"
        pi = map_pi(str(row[12]).strip() if row[12] else "")
        prepayment = float(row[13]) if row[13] else 0
        payment_method = str(row[16]).strip() if row[16] else None
        client_debt = float(row[17]) if row[17] else 0
        partner = str(row[23]).strip() if len(row) > 23 and row[23] else None
        notes = str(row[25]).strip() if len(row) > 25 and row[25] else None

        # Get or create client
        if client_name not in client_cache:
            client_cache[client_name] = create_client(client_name)
        client_id = client_cache[client_name]

        # Create product
        product_id = create_product(trademark, weaving, layer, doc_type, tr_ts)

        # Determine price based on doc_type
        if doc_type == "CC":
            total_price = 30000
        else:
            total_price = 25000

        # Create order
        order_payload = {
            "client_id": client_id,
            "doc_type": doc_type,
            "tr_ts": tr_ts,
            "duration_years": 1,
            "protocol_count": 1,
            "status": status,
            "layout_status": layout,
            "sample_status": sample,
            "cert_body": cert_body,
            "original_status": original if original else "Нет",
            "pi_status": pi,
            "expected_date": expected_date,
            "prepayment": prepayment,
            "payment_method": payment_method,
            "partner": partner,
            "notes": notes,
            "product_ids": [product_id],
        }

        r = requests.post(f"{API}/orders/", json=order_payload)
        if r.status_code == 200:
            oid = r.json()["id"]
            # Update actual_date and client_debt via PATCH
            patch = {}
            if actual_date:
                patch["actual_date"] = actual_date
            if client_debt:
                patch["client_debt"] = client_debt
            if patch:
                requests.patch(f"{API}/orders/{oid}", json=patch)
            created += 1
            print(f"  [{created}] {brand_raw[:50]} -> {status}")
        else:
            print(f"  FAIL: {brand_raw[:50]} -> {r.status_code}: {r.text[:100]}")

    print(f"\nDone! Created {created} orders from {len(client_cache)} clients.")


if __name__ == "__main__":
    main()
